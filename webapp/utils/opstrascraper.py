
import csv
from xmlrpc.client import DateTime
import openpyxl
import pandas as pd
from datetime import date,datetime
import time
import json
import requests
import random
import os

from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions

from webapp.utils.apimanager import ApiManager


apimgr = ApiManager()
    
class OpstraScraper():
    
    def getFilePath(self):
        if("E:" in os.getcwd()):
            #return "/Users/noaman/Documents/DATA/dev/code/Django/optionstrader/webapp/utils/"
            #return "/Users/noam/Documents/dev/code/django/optionstrader/webapp/utils/"
            return "/optionstrader/webapp/utils/"
        else:
            return "/var/www/optionstrader/webapp/utils/"
    
   
    
    
    def write_to_CSV(self, filename, data):
        
        with open(  filename, 'w', newline='') as csvfile:
            fieldnames = data[0].keys()
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames , lineterminator='\n')

            writer.writeheader()
            writer.writerows(data)

    def getDriver(self) :        #wdPath = 'webdriver/msedgedriver.exe'):
        webdriver_path =  'webdriver/msedgedriver.exe'


        # Create Edge WebDriver with headless option
        #edge_service = EdgeService(webdriver_path)
        edge_options = EdgeOptions()
        edge_options.use_chromium = True  # Use Chromium-based Edge
        edge_options.headless = True



        # Initialize the Selenium WebDriver 
        driver = webdriver.Chrome(executable_path= webdriver_path , options=edge_options)  #    .Chrome(executable_path='E:/chrome-win32/chrome.exe')
        #driver = webdriver.Firefox(executable_path='C:/Program Files/Mozilla Firefox/firefox.exe')

            
        return driver      

    def  getDataFromLink(self,driver, link, dict_name) : 
        # Navigate to the desired URL
        driver.get(link)  # json data

        # Wait for some time to allow the page to load 
        time.sleep(5)

        page_source = driver.page_source
        # Parse the HTML string with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find the <div> element with id "json-data"
        json_body = soup.find('body')
        json_div = json_body.find('div', {'hidden': 'true'})


        # Check if the <div> element exists and has content
        if json_div and json_div.string:
            # Extract the JSON data
            json_data = json_div.string

            # Parse the JSON data into a Python dictionary
            json_object = json.loads(json_data)[dict_name]

            # Now, 'json_object' contains the extracted JSON data
            #print(json_object)
        else:
            print("JSON data not found.")


        return json_object


    def login(self):
        driver = self.getDriver()
        # Open the website
        #print(driver.current_url)

        ####################      login              ###############################################
        driver.get("https://sso.definedge.com/auth/realms/definedge/protocol/openid-connect/auth?response_type=code&client_id=opstra&redirect_uri=https://opstra.definedge.com/ssologin&state=e2cf559f-356c-425a-87e3-032097f643d0&login=true&scope=openid")

        time.sleep(5)

        # Find the username and password fields
        username_field = driver.find_element_by_id("username")
        password_field = driver.find_element_by_id("password")

        #  fill the username and password fields
        username_field.send_keys("zainab.kapadia@gmail.com")
        password_field.send_keys("Idontknow0811")

        # Find and click the login button
        login_button = driver.find_element_by_id("kc-login")
        login_button.click()

        # Wait for some time to allow the login process to complete (adjust the time as needed)
        time.sleep(2)

        return driver


    def writeOptionsData(self):
        link = "https://opstra.definedge.com/api/optionsdashboard/free"
        
        driver = self.login()
        
        json_data = self.getDataFromLink(driver, link, 'optiondata')
        
        driver.close()
        
        # save it to a JSON file if needed
        with open("options.json", "w") as json_file:
            json.dump(json_data, json_file, indent=4)

        csv_file_path = "Output/opstra_options.csv"

        self.write_to_CSV(csv_file_path, json_data)   
        
        
    def  getOpstraData(self, driver, link, dict_name) : 
        # Navigate to the desired URL
        driver.get(link)  # json data

        # Wait for some time to allow the page to load 
        time.sleep(5)

        page_source = driver.page_source
        # Parse the HTML string with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find the <div> element with id "json-data"
        json_body = soup.find('body')
        json_div = json_body.find('div', {'hidden': 'true'})


        # Check if the <div> element exists and has content
        if json_div and json_div.string:
            # Extract the JSON data
            json_data = json_div.string

            # Parse the JSON data into a Python dictionary
            json_object = json.loads(json_data)[dict_name]

            # Now, 'json_object' contains the extracted JSON data
            #print(json_object)
            return json_object
        else:
            print("JSON data not found.")
            return  []
    
    
    def get_symbol_data(self,driver , symbol = "COALINDIA"):
        #   link  = https://opstra.definedge.com/api/openinterest/futuresopeninterest/COALINDIA&Combined%20OpenInterest
        #   COALINDIA  shouldbe a variable 

        link = "https://opstra.definedge.com/api/openinterest/futuresopeninterest/" + symbol + "&Combined%20OpenInterest"

        data_list = self.getOpstraData(driver, link, 'data')


        ''' # save it to a JSON file if needed
        json_filepath = symbol + "_futuresOI.json"
        with open( json_filepath , "w") as json_file:
            json.dump(data_list, json_file, indent=4)
            '''
            
            
            
        #   Convert the list to a dictionary
        # Define your headers as a list
        headers = ['0',	'Open',	'High'	,'Low',	'Close',  'OI', 'Futures_Vol',	'BuildUp',	'8', '9',  'Cash Delivery', 'CashDelofVolume']


        last_21_rows = data_list[-21:]

        # Create a dictionary using zip
        data_dict = [dict(zip(headers, row)) for row in last_21_rows]

        

        return data_dict

    def writeFuturesData(self):
        link = "https://opstra.definedge.com/api/openinterest/futuresbuildup"
        
        driver = self.login()
        data_dict = self.getDataFromLink(driver, link, 'data')
        
        driver.close()

        # save it to a JSON file if needed
        with open("futures.json", "w") as json_file:
            json.dump(data_dict, json_file, indent=4)

        csv_file_path = "Output/opstra_futures.csv"

        self.write_to_CSV( csv_file_path, data_dict) 
        
    
    def getDataFrame(self ,driver,  symbol = 'COALINDIA'):
        data_dict = self.get_symbol_data(driver, symbol)

        df = pd.DataFrame(data_dict)
        df['SpotChange'] = df['Close'] - df['Close'].shift(1)
        df['Vol_Change'] = df['Futures_Vol'] - df['Futures_Vol'].shift(1)
        df['Vol_Change%'] = ( df['Vol_Change']/df['Futures_Vol'].shift(1) )* 100
        df['OI_Change'] = df['OI'] - df['OI'].shift(1)
        #df['LongBuildup'] = df['BuildUp'].apply(lambda x: df['OI'] if (x =='green' or x=='lightgreen') else 0 )
    
        condition_G_VOL = df['BuildUp'].str.contains('green' , case=False) & (df['Vol_Change%'] > 0)
        df.loc[condition_G_VOL, 'LongBuildup'] = df.loc[condition_G_VOL , 'OI']
        #df.loc[df['BuildUp'] =='green' , 'LongBuildup'] = df.loc[df['BuildUp'] =='green' , 'OI']
        #df.loc[df['BuildUp'] =='lightgreen' , 'LongBuildup'] =  df.loc[df['BuildUp'] =='lightgreen' , 'OI']

        condition_RO_VOL = (df['BuildUp'].str.contains('red' , case=False) | df['BuildUp'].str.contains('orange' , case=False) ) &  (df['Vol_Change%'] > 0)
        df.loc[condition_RO_VOL , 'ShortBuildup'] = df.loc[condition_RO_VOL , 'OI']

        df.fillna(value='0', inplace=True)

        #print(df)
        #df.to_csv("COALINDIA.csv")

        return df
    
    
    def writeAnalysisData(self , symbol_string):
        link  = ""
        driver = self.login()
        #symbol_string =    "SUNPHARMA, ASIANPAINT, INDUSINDBK, JUBLFOOD, IPCALAB, AXISBANK, INDIAMART, CANBK, PFC"
        
        symbol_list = [    item.strip() for item in symbol_string.split(',')   ]

        output_list = []

        for symbol in symbol_list:
            df = self.getDataFrame(driver,symbol)
        
            stockname = symbol
            link = "https://opstra.definedge.com/api/openinterest/futuresopeninterest/" + symbol + "&Combined%20OpenInterest"
            
            
            cum_SpotChange =  df['SpotChange'].astype(float).sum()
            cum_Vol_Change = df['Vol_Change'].astype(float).sum()
            cum_Vol_Change_percent = df['Vol_Change%'].astype(float).sum()
            cum_OI_Change = df['OI_Change'].astype(float).sum()
            cum_LongBuildup = df['LongBuildup'].astype(float).sum()
            cum_ShortBuildup = df['ShortBuildup'].astype(float).sum()
            cum_OI = df['OI'].astype(float).sum()
            
            cum_BU_Change = cum_LongBuildup - cum_ShortBuildup
            cum_BU_Change_percent = (cum_LongBuildup + cum_ShortBuildup)/cum_OI * 100
            
            
            excel_filename =  "Output/" +  stockname + "_futuresOI.xlsx"
            df.to_excel(excel_filename)
            wb = openpyxl.load_workbook(excel_filename)
            ws = wb['Sheet1']
            ws.insert_rows(1, amount=5)
            
            ws.cell(2,1).value = 'STOCKNAME'
            ws.cell(2,2).value = stockname
            ws.cell(2,4 ).value = link
            
            ws.cell(4,1).value = 'CumBUChange'
            ws.cell(4,2).value = cum_BU_Change
            
            ws.cell(3,1).value = 'RESULT'
            result = ''
            if cum_BU_Change < 0 :
                result  = 'BEARISH'
            else:
                result  = 'BULLISH'
            ws.cell(3,2).value = result    
            
            ws.cell(5,1).value = 'CumBUChange%'
            ws.cell(5,2).value = cum_BU_Change_percent
            
            ws.cell(14,5).value = cum_SpotChange
            
            
            ws.cell(5, 15).value = cum_Vol_Change
            ws.cell(5,16).value = cum_Vol_Change_percent
            ws.cell(5,17).value = cum_OI_Change
            ws.cell(5,18).value = cum_LongBuildup
            ws.cell(5,19).value = cum_ShortBuildup
            ws.cell(5,7).value = cum_OI
            
            
            wb.save(excel_filename)
            
            output_list.append([stockname, datetime.now().strftime("%Y-%m-%d_%H:%M:%S") , result,  cum_BU_Change, cum_BU_Change_percent])
            
                
        # Close the WebDriver
        driver.quit()



        output_df = pd.DataFrame(output_list)
        output_df.columns = ['STOCKNAME', 'DATE', 'RESULT', 'CumBUChange', 'CumBUChange%']
        output_df.to_excel("Output/Opstra_Data_Screener_Analysis.xlsx")   

                
        
                
         

if __name__ == '__main__':
    scraper = OpstraScraper()
    #scraper.writeOptionsData()
    #scraper.writeFuturesData()
    scraper.writeAnalysisData()
